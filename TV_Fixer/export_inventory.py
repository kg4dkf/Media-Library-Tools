#!/usr/bin/env python3
"""
Export a phone-friendly inventory of your TV collection. Cross-references
the series_stats_detail.xlsx (expected episodes per season) against what
you actually have on disk, and produces:

  - inventory.csv    flat per-show table (status, have/expected, missing list)
  - inventory.html   simple printable HTML that renders well on a phone

Usage:
    python export_inventory.py "G:\\TV" ^
        --xlsx "H:\\_mediaprep\\series_stats_detail.xlsx" ^
        --out inventory.csv --html inventory.html

Status per show:
    complete       you have every expected regular episode
    near_complete  >= 90% of expected
    partial        between 10% and 90%
    barely_started < 10%
    missing        0 episodes on disk
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple


def _safe(s) -> str:
    """Round-trip strings through UTF-8 with errors='replace' so unpaired
    Windows surrogates don't crash csv/html writers."""
    return str(s).encode("utf-8", errors="replace").decode("utf-8")

VIDEO_EXTS = {
    ".mkv", ".mp4", ".avi", ".mov", ".m4v", ".ts", ".m2ts",
    ".mpg", ".mpeg", ".wmv", ".webm",
}

TMDB_FOLDER_RE = re.compile(r"\{tmdb-(\d+)\}", re.IGNORECASE)
SE_RE = re.compile(r"[Ss](\d{1,2})[\s._-]*[Ee](\d{1,3})")
SEASON_FOLDER_RE = re.compile(r"^[Ss]eason\s+(\d{1,2})$")


def scan_disk(media_root: Path) -> Dict[str, Dict[int, set]]:
    """For each show folder under media_root, collect the set of episode
    numbers found per season. Returns {tmdb_id: {season: {ep, ep, ...}}}
    keyed by TMDB id when the folder has a {tmdb-NNNN} marker; otherwise
    keyed by the folder name itself."""
    inv: Dict[str, Dict[int, set]] = defaultdict(lambda: defaultdict(set))
    for show_dir in sorted(media_root.iterdir()):
        if not show_dir.is_dir():
            continue
        m = TMDB_FOLDER_RE.search(show_dir.name)
        key = m.group(1) if m else show_dir.name
        for path in show_dir.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in VIDEO_EXTS:
                continue
            mse = SE_RE.search(path.name)
            if mse:
                season, ep = int(mse.group(1)), int(mse.group(2))
                inv[key][season].add(ep)
                continue
            # Fall back: parent folder is "Season N" + filename has episode number
            parent = path.parent
            mseason = SEASON_FOLDER_RE.match(parent.name)
            if mseason:
                season = int(mseason.group(1))
                # Try to pluck a stray episode number from the filename
                mep = re.search(r"(?:^|[^0-9])(\d{1,3})(?:[^0-9]|$)", path.name)
                if mep:
                    inv[key][season].add(int(mep.group(1)))
    return inv


def load_xlsx(xlsx: Path, summary_sheet: str, detail_sheet: str
              ) -> Tuple[List[dict], Dict[str, Dict[int, int]]]:
    """Returns (summary_rows, expected_by_season)
    expected_by_season: {tmdb_id_str: {season: expected_ep_count}}"""
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")
    wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
    summary = wb[summary_sheet]
    detail = wb[detail_sheet]

    # Summary -> list of show dicts
    rows = summary.iter_rows(values_only=True)
    s_header = [str(h).strip() if h is not None else "" for h in next(rows)]
    summary_rows = [dict(zip(s_header, r)) for r in rows if r and r[0] is not None]

    # Detail -> expected episodes per (tmdb_id, season)
    rows = detail.iter_rows(values_only=True)
    d_header = [str(h).strip() if h is not None else "" for h in next(rows)]
    di = {h: i for i, h in enumerate(d_header)}
    expected: Dict[str, Dict[int, int]] = defaultdict(dict)
    for r in rows:
        if not r or r[di["tmdb_id"]] is None:
            continue
        tmdb_id = str(r[di["tmdb_id"]])
        season = r[di["season"]]
        ep_expected = r[di["episodes_expected"]]
        if season is None or ep_expected is None:
            continue
        expected[tmdb_id][int(season)] = int(ep_expected)
    return summary_rows, expected


def status_for(have_total: int, expected_total: int) -> str:
    if expected_total == 0:
        return "no_expected"
    if have_total == 0:
        return "missing"
    pct = have_total / expected_total
    if pct >= 1.0:
        return "complete"
    if pct >= 0.9:
        return "near_complete"
    if pct >= 0.1:
        return "partial"
    return "barely_started"


def compact_missing(have: set, expected: int) -> str:
    """Return a compact summary like '3-5, 8, 12-22' of missing episodes."""
    if expected == 0:
        return ""
    miss = sorted(set(range(1, expected + 1)) - have)
    if not miss:
        return ""
    runs: List[Tuple[int, int]] = []
    start = prev = miss[0]
    for n in miss[1:]:
        if n == prev + 1:
            prev = n
        else:
            runs.append((start, prev))
            start = prev = n
    runs.append((start, prev))
    return ", ".join(f"{a}" if a == b else f"{a}-{b}" for a, b in runs)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("media_root", type=Path)
    ap.add_argument("--xlsx", type=Path, required=True,
                    help="series_stats_detail.xlsx")
    ap.add_argument("--summary-sheet", default="series_stats_summary")
    ap.add_argument("--detail-sheet", default="series_stats_detail")
    ap.add_argument("--out", type=Path, default=Path("inventory.csv"))
    ap.add_argument("--html", type=Path, default=Path("inventory.html"))
    ap.add_argument("--sort", default="status",
                    choices=["status", "completeness", "title", "missing"],
                    help="Output sort order")
    args = ap.parse_args()

    if not args.media_root.is_dir():
        sys.exit(f"Not a directory: {args.media_root}")

    inv = scan_disk(args.media_root)
    summary_rows, expected_by_season = load_xlsx(
        args.xlsx, args.summary_sheet, args.detail_sheet,
    )

    out_rows = []
    for s in summary_rows:
        tmdb_id = str(s.get("tmdb_id"))
        title = s.get("title") or ""
        year = s.get("year") or ""
        on_disk = inv.get(tmdb_id, {})
        seasons_expected = expected_by_season.get(tmdb_id, {})
        have_total = sum(len(on_disk.get(season, set()))
                         for season in seasons_expected)
        expected_total = sum(seasons_expected.values())
        status = status_for(have_total, expected_total)
        # Compact missing summary across all seasons.
        miss_parts = []
        for season in sorted(seasons_expected.keys()):
            mp = compact_missing(on_disk.get(season, set()),
                                 seasons_expected[season])
            if mp:
                miss_parts.append(f"S{season:02d}: {mp}")
        out_rows.append({
            "status": status,
            "title": title,
            "year": year,
            "have": have_total,
            "expected": expected_total,
            "pct": (have_total / expected_total * 100) if expected_total else 0,
            "missing_summary": " | ".join(miss_parts),
            "tmdb_id": tmdb_id,
        })

    # Sort
    if args.sort == "status":
        # Show missing & barely_started first (most actionable for thrift store)
        order = {"missing": 0, "barely_started": 1, "partial": 2,
                 "near_complete": 3, "complete": 4, "no_expected": 5}
        out_rows.sort(key=lambda r: (order.get(r["status"], 9), r["title"].lower()))
    elif args.sort == "completeness":
        out_rows.sort(key=lambda r: (r["pct"], r["title"].lower()))
    elif args.sort == "missing":
        out_rows.sort(key=lambda r: (-(r["expected"] - r["have"]), r["title"].lower()))
    else:
        out_rows.sort(key=lambda r: r["title"].lower())

    # CSV
    with args.out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "status", "title", "year", "have", "expected", "pct",
            "missing_summary", "tmdb_id",
        ])
        w.writeheader()
        for r in out_rows:
            r2 = {k: (_safe(v) if isinstance(v, str) else v)
                  for k, v in r.items()}
            r2["pct"] = f"{r['pct']:.1f}"
            w.writerow(r2)

    # HTML (phone-friendly)
    if args.html:
        rows_html = []
        for r in out_rows:
            color = {
                "missing":        "#ffd6d6",
                "barely_started": "#ffe6c2",
                "partial":        "#fff2b3",
                "near_complete":  "#d9f2d9",
                "complete":       "#b3e6b3",
                "no_expected":    "#eeeeee",
            }.get(r["status"], "#ffffff")
            rows_html.append(
                f"<tr style='background:{color};'>"
                f"<td>{r['status']}</td>"
                f"<td>{_safe(r['title'])} ({r['year']})</td>"
                f"<td>{r['have']}/{r['expected']}</td>"
                f"<td>{r['pct']:.0f}%</td>"
                f"<td>{_safe(r['missing_summary'])}</td>"
                f"</tr>"
            )
        html = f"""<!doctype html><html><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<title>TV inventory</title>
<style>
body {{ font-family: -apple-system, system-ui, sans-serif; margin: 8px; font-size: 14px; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ padding: 4px 6px; border: 1px solid #ccc; vertical-align: top; }}
th {{ position: sticky; top: 0; background: #fff; }}
input {{ width: 100%; padding: 8px; font-size: 16px; box-sizing: border-box; }}
</style></head><body>
<input id='q' placeholder='Filter by title...' oninput="
  var v = this.value.toLowerCase();
  document.querySelectorAll('tbody tr').forEach(function(r) {{
    r.style.display = r.cells[1].textContent.toLowerCase().includes(v) ? '' : 'none';
  }});
">
<table>
<thead><tr><th>Status</th><th>Show</th><th>Have</th><th>%</th><th>Missing</th></tr></thead>
<tbody>
{chr(10).join(rows_html)}
</tbody></table>
</body></html>
"""
        args.html.write_text(html, encoding="utf-8")

    counts = defaultdict(int)
    for r in out_rows:
        counts[r["status"]] += 1
    print(f"Inventory: {len(out_rows)} shows")
    for st in ["complete", "near_complete", "partial", "barely_started", "missing", "no_expected"]:
        if counts[st]:
            print(f"  {st:15s}  {counts[st]}")
    print(f"  CSV:  {args.out}")
    if args.html:
        print(f"  HTML: {args.html}")



if __name__ == "__main__":
    main()
