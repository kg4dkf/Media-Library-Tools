#!/usr/bin/env python3
"""
Download English subtitles from OpenSubtitles, for one show or for a whole
list of shows (xlsx / csv / txt).

Requires OPENSUBTITLES_API_KEY env var (set up at opensubtitles.com).
Output layout (matches what index_subs.py expects):
    <out>/<show>/SXXEYY.srt

Rate-limit note:
- Free tier: 5 downloads/day.
- VIP ($3/mo): 1000 downloads/day.
- 'Listing' (asking the API which subtitles exist for a show) is NOT rate-limited
  in the same way, so this script parallelizes the listing step and serializes
  downloads. When the daily download quota is reached, the script stops cleanly
  and you can resume tomorrow -- already-downloaded SRTs are skipped.

Single-show usage:
    python fetch_subtitles.py --show "Avatar: The Last Airbender" --out subs
    python fetch_subtitles.py --tmdb 246 --out subs
    python fetch_subtitles.py --imdb 0417299 --out subs

Batch usage (xlsx):
    python fetch_subtitles.py --from-xlsx series_stats_detail.xlsx ^
        --sheet series_stats_summary ^
        --title-col title --tmdb-col tmdb_id ^
        --out subs --workers 6 --max-total-downloads 900

Batch usage (csv or txt):
    python fetch_subtitles.py --from-csv shows.csv ^
        --title-col title --tmdb-col tmdb_id --out subs
    python fetch_subtitles.py --from-txt shows.txt --out subs
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import requests
from tqdm import tqdm

API = "https://api.opensubtitles.com/api/v1"
UA = "tv-id-pipeline/1.0"


class QuotaExhausted(Exception):
    """Raised when OpenSubtitles signals daily download quota exceeded."""


def make_session(api_key: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Api-Key": api_key,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": UA,
    })
    return s


def search_show(s: requests.Session, name: str) -> Optional[dict]:
    """Title search fallback when we have neither TMDB nor IMDb id."""
    r = s.get(f"{API}/features",
              params={"query": name, "type": "tvshow"}, timeout=30)
    r.raise_for_status()
    data = r.json().get("data", [])
    for hit in data:
        if hit.get("attributes", {}).get("feature_type") == "Tvshow":
            return hit
    return data[0] if data else None


def list_subtitles(
    s: requests.Session,
    *,
    parent_tmdb_id: Optional[str] = None,
    parent_imdb_id: Optional[str] = None,
    language: str = "en",
    season: Optional[int] = None,
) -> List[dict]:
    """List subtitles for all episodes of a show. Paginated.

    We do NOT pass `order_by` -- the API rejects unsupported values with
    HTTP 400 on big shows (Friends, Family Guy, etc). We sort client-side
    in pick_best_per_episode() instead.

    If we hit a 4xx mid-pagination, we return what we already collected
    rather than aborting -- a partial list is still useful.
    """
    if not (parent_tmdb_id or parent_imdb_id):
        return []
    results: List[dict] = []
    page = 1
    while True:
        params = {
            "languages": language,
            "type": "episode",
            "page": str(page),
        }
        if parent_tmdb_id:
            params["parent_tmdb_id"] = str(parent_tmdb_id)
        if parent_imdb_id:
            params["parent_imdb_id"] = str(parent_imdb_id)
        if season is not None:
            params["season_number"] = str(season)

        r = s.get(f"{API}/subtitles", params=params, timeout=30)
        if r.status_code == 429:
            time.sleep(2.0)
            continue
        if 400 <= r.status_code < 500:
            # Permanent client error on this page -- keep what we have.
            if not results:
                r.raise_for_status()
            break
        r.raise_for_status()
        body = r.json()
        results.extend(body.get("data", []))
        total_pages = body.get("total_pages", 1) or 1
        if page >= total_pages:
            break
        page += 1
        time.sleep(0.15)
    return results


def pick_best_per_episode(subs: Iterable[dict]) -> List[dict]:
    """For each (season, episode), keep the most-downloaded subtitle entry."""
    best: dict[Tuple[int, int], dict] = {}
    for sub in subs:
        attrs = sub.get("attributes", {})
        feat = attrs.get("feature_details", {}) or {}
        season = feat.get("season_number")
        episode = feat.get("episode_number")
        if season is None or episode is None:
            continue
        key = (int(season), int(episode))
        dl = attrs.get("download_count", 0) or 0
        cur = best.get(key)
        if cur is None or (cur.get("attributes", {}).get("download_count", 0) or 0) < dl:
            best[key] = sub
    return [best[k] for k in sorted(best.keys())]


def download_one(s: requests.Session, file_id: int) -> bytes:
    """Download one subtitle. Raises QuotaExhausted if daily quota hit."""
    r = s.post(f"{API}/download", json={"file_id": file_id}, timeout=30)
    if r.status_code == 406:
        raise QuotaExhausted(r.text[:200])
    if r.status_code == 429:
        time.sleep(3.0)
        r = s.post(f"{API}/download", json={"file_id": file_id}, timeout=30)
        if r.status_code == 406:
            raise QuotaExhausted(r.text[:200])
    if r.status_code != 200:
        raise RuntimeError(f"download request failed: {r.status_code} {r.text[:200]}")
    link = r.json().get("link")
    if not link:
        raise RuntimeError("download response had no link")
    dl = s.get(link, timeout=60)
    dl.raise_for_status()
    return dl.content


def safe_show_dir(name: str) -> str:
    out = re.sub(r"[^A-Za-z0-9 ._-]", "", name).strip()
    return re.sub(r"\s+", " ", out) or "show"


# Plex / Sonarr embed TMDB ids in folder names like:
#   "Friends (1994) {tmdb-1668}"
#   "Avatar - The Last Airbender (2005) {tmdb-246}"
_TMDB_FOLDER_RE = re.compile(r"\{tmdb-(\d+)\}", re.IGNORECASE)

# Leading articles we strip during name-based normalization to absorb the
# common 'The X' vs 'X' folder/title difference.
_LEADING_ARTICLES = ("the ", "a ", "an ")


def _strip_leading_article(s: str) -> str:
    low = s.lower().lstrip()
    for art in _LEADING_ARTICLES:
        if low.startswith(art):
            return s.lstrip()[len(art):]
    return s


def _norm_for_match(s: str) -> str:
    """Aggressive normalization for fuzzy folder<->show matching.

    Strip a leading article, lowercase, drop everything that isn't a letter
    or digit. So 'The Adventures of Superman (1952) {tmdb-12662}' and
    'Adventures of Superman' both collapse to 'adventuresofsuperman' for
    the name-prefix path (TMDB id matching, when available, is preferred).
    """
    s2 = _strip_leading_article(s)
    return re.sub(r"[^a-z0-9]+", "", s2.lower())


def build_folder_index(media_root: Path) -> dict:
    """One-shot scan of media_root's immediate children.

    Returns a dict with two keys:
      'by_tmdb': {tmdb_id_str: [Path, ...]}
      'by_name': {normalized_name: [Path, ...]}
    The TMDB index is built from {tmdb-NNNN} markers embedded in folder
    names; the name index is the fuzzy fallback.
    """
    idx_tmdb: dict[str, List[Path]] = {}
    idx_name: dict[str, List[Path]] = {}
    if not media_root or not media_root.is_dir():
        return {"by_tmdb": idx_tmdb, "by_name": idx_name}
    for child in media_root.iterdir():
        if not child.is_dir():
            continue
        m = _TMDB_FOLDER_RE.search(child.name)
        if m:
            idx_tmdb.setdefault(m.group(1), []).append(child)
        idx_name.setdefault(_norm_for_match(child.name), []).append(child)
    return {"by_tmdb": idx_tmdb, "by_name": idx_name}


def find_show_folder(folder_idx: dict, show_label: str,
                     tmdb_id: Optional[str] = None) -> Optional[Path]:
    """Match a show against the folder index.

    Match order:
      1. TMDB id (perfect match: folder contains {tmdb-NNNN} matching the
         show's TMDB id).
      2. Exact normalized name match.
      3. Folders whose normalized name starts with the show's normalized
         name, or vice versa.

    Among multiple candidates, picks the folder whose original name length
    is closest to the show label's length (heuristic for "most specific").
    """
    by_tmdb = folder_idx.get("by_tmdb", {})
    by_name = folder_idx.get("by_name", {})

    if tmdb_id:
        tid = str(tmdb_id).strip()
        if tid in by_tmdb:
            paths = by_tmdb[tid]
            paths.sort(key=lambda p: abs(len(p.name) - len(show_label)))
            return paths[0]

    key = _norm_for_match(show_label)
    if not key:
        return None

    if key in by_name:
        cands = list(by_name[key])
    else:
        cands = []
        for k, paths in by_name.items():
            if k.startswith(key) or key.startswith(k):
                cands.extend(paths)
        if not cands:
            return None

    cands.sort(key=lambda p: abs(len(p.name) - len(show_label)))
    return cands[0]


@dataclass
class ShowSpec:
    label: str
    tmdb_id: Optional[str] = None
    imdb_id: Optional[str] = None
    query: Optional[str] = None


def _col_idx(header: List[str], name: Optional[str]) -> Optional[int]:
    if not name:
        return None
    if name in header:
        return header.index(name)
    for i, h in enumerate(header):
        if h.lower() == name.lower():
            return i
    return None


def load_xlsx(path: Path, sheet: Optional[str], title_col: str,
              tmdb_col: Optional[str], imdb_col: Optional[str]) -> List[ShowSpec]:
    try:
        import openpyxl
    except Exception:
        sys.exit("pip install openpyxl to read .xlsx")
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header = [str(h).strip() if h is not None else "" for h in header]
    ti = _col_idx(header, title_col)
    ki = _col_idx(header, tmdb_col) if tmdb_col else None
    ii = _col_idx(header, imdb_col) if imdb_col else None
    if ti is None:
        sys.exit(f"Title column {title_col!r} not in sheet header: {header}")
    out: List[ShowSpec] = []
    for row in rows:
        if not row:
            continue
        title = row[ti] if ti < len(row) else None
        if not title:
            continue
        tmdb = None
        imdb = None
        if ki is not None and ki < len(row) and row[ki] not in (None, ""):
            tmdb = str(row[ki]).strip()
        if ii is not None and ii < len(row) and row[ii] not in (None, ""):
            imdb = str(row[ii]).strip()
        out.append(ShowSpec(label=str(title).strip(), tmdb_id=tmdb,
                            imdb_id=imdb, query=str(title).strip()))
    return out


def load_csv(path: Path, title_col: str, tmdb_col: Optional[str],
             imdb_col: Optional[str]) -> List[ShowSpec]:
    out: List[ShowSpec] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get(title_col) or "").strip()
            if not title:
                continue
            tmdb = (row.get(tmdb_col) or "").strip() if tmdb_col else ""
            imdb = (row.get(imdb_col) or "").strip() if imdb_col else ""
            out.append(ShowSpec(label=title, query=title,
                                tmdb_id=tmdb or None, imdb_id=imdb or None))
    return out


def load_txt(path: Path) -> List[ShowSpec]:
    out: List[ShowSpec] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        out.append(ShowSpec(label=line, query=line))
    return out


@dataclass
class ListedShow:
    spec: ShowSpec
    picked: List[dict]
    error: Optional[str] = None


def resolve_and_list(s: requests.Session, spec: ShowSpec,
                     language: str, season: Optional[int]) -> ListedShow:
    """Resolve identifiers and list available subs for one show."""
    try:
        tmdb = spec.tmdb_id
        imdb = spec.imdb_id
        if not tmdb and not imdb and spec.query:
            hit = search_show(s, spec.query)
            if hit:
                attrs = hit.get("attributes", {})
                imdb = str(attrs.get("imdb_id") or "").lstrip("0") or None
                tmdb_from_search = attrs.get("tmdb_id")
                if tmdb_from_search:
                    tmdb = str(tmdb_from_search)
        if not tmdb and not imdb:
            return ListedShow(spec=spec, picked=[], error="no_id")
        subs = list_subtitles(s, parent_tmdb_id=tmdb, parent_imdb_id=imdb,
                              language=language, season=season)
        picked = pick_best_per_episode(subs)
        return ListedShow(spec=spec, picked=picked)
    except requests.HTTPError as e:
        code = getattr(e.response, "status_code", "?")
        return ListedShow(spec=spec, picked=[], error=f"http_{code}")
    except Exception as e:
        return ListedShow(spec=spec, picked=[], error=f"err:{e}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    src = ap.add_argument_group("input (pick one or combine)")
    src.add_argument("--show", help="Single show: title")
    src.add_argument("--imdb", help="Single show: IMDb parent id digits")
    src.add_argument("--tmdb", help="Single show: TMDB parent id")
    src.add_argument("--from-xlsx", type=Path, help="Excel file with shows list")
    src.add_argument("--from-csv", type=Path, help="CSV file with shows list")
    src.add_argument("--from-txt", type=Path, help="Plain-text file, one show per line")

    batch = ap.add_argument_group("batch-input columns (xlsx/csv)")
    batch.add_argument("--sheet", help="xlsx sheet name (default = first sheet)")
    batch.add_argument("--title-col", default="title", help="Column with show title")
    batch.add_argument("--tmdb-col", default="tmdb_id",
                       help="Column with TMDB id (preferred). Pass '' to disable.")
    batch.add_argument("--imdb-col", default=None,
                       help="Column with IMDb id (optional)")

    ap.add_argument("--language", default="en")
    ap.add_argument("--season", type=int, help="Only fetch this season (single-show mode)")
    ap.add_argument("--out", type=Path, default=Path("subs"),
                    help="Fallback output root when a show folder cannot be matched under --media-root")
    ap.add_argument("--media-root", type=Path, default=None,
                    help="Root of your TV collection (e.g. G:\\TV). When set, "
                         "subtitles are saved directly into the matched show "
                         "folder so Plex can pick them up after rename.")
    ap.add_argument("--save-mapping", type=Path, default=None,
                    help="Write a CSV of (show, matched_folder) so you can review the matches")
    ap.add_argument("--workers", type=int, default=6,
                    help="Concurrent threads for the LISTING step. Downloads stay serial.")
    ap.add_argument("--max-per-show", type=int, default=0,
                    help="Cap downloads per show (0 = unlimited)")
    ap.add_argument("--max-total-downloads", type=int, default=0,
                    help="Hard cap on total downloads this run (0 = no cap)")
    ap.add_argument("--rate-delay", type=float, default=0.4,
                    help="Seconds to sleep between download API calls")
    ap.add_argument("--dry-run", action="store_true",
                    help="List what would be downloaded; no downloads")
    ap.add_argument("--skip-complete", action="store_true", default=True,
                    help="Skip shows whose output folder already has >= listed sub count")
    ap.add_argument("--no-skip-complete", action="store_false", dest="skip_complete")
    args = ap.parse_args()

    api_key = os.environ.get("OPENSUBTITLES_API_KEY")
    if not api_key:
        sys.exit("Set OPENSUBTITLES_API_KEY env var.")

    specs: List[ShowSpec] = []
    if args.show or args.imdb or args.tmdb:
        specs.append(ShowSpec(
            label=args.show or (f"tmdb-{args.tmdb}" if args.tmdb else f"imdb-{args.imdb}"),
            tmdb_id=args.tmdb, imdb_id=args.imdb, query=args.show))
    if args.from_xlsx:
        specs.extend(load_xlsx(args.from_xlsx, args.sheet, args.title_col,
                               args.tmdb_col or None, args.imdb_col))
    if args.from_csv:
        specs.extend(load_csv(args.from_csv, args.title_col,
                              args.tmdb_col or None, args.imdb_col))
    if args.from_txt:
        specs.extend(load_txt(args.from_txt))

    if not specs:
        sys.exit("No input. Pass --show / --imdb / --tmdb, "
                 "or --from-xlsx / --from-csv / --from-txt.")

    print(f"{len(specs)} show(s) to process.")

    listed: List[ListedShow] = []

    def worker(spec: ShowSpec) -> ListedShow:
        sess = make_session(api_key)
        return resolve_and_list(sess, spec, args.language, args.season)

    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futures = [ex.submit(worker, spec) for spec in specs]
        with tqdm(total=len(futures), desc="listing", unit="show") as bar:
            for fut in as_completed(futures):
                listed.append(fut.result())
                bar.update(1)

    listed.sort(key=lambda x: x.spec.label.lower())
    total_listed = sum(len(x.picked) for x in listed)
    errs = [x for x in listed if x.error]
    print(f"Listing done: {total_listed} subs across "
          f"{len(listed) - len(errs)} resolvable shows. "
          f"{len(errs)} unresolvable.")
    for x in errs[:10]:
        print(f"  ! {x.spec.label}: {x.error}")
    if len(errs) > 10:
        print(f"  ... and {len(errs) - 10} more")

    # Build a folder index for --media-root if provided.
    folder_idx = build_folder_index(args.media_root) if args.media_root else {}
    if args.media_root:
        print(f"Indexed {sum(len(v) for v in folder_idx.values())} folder(s) "
              f"under {args.media_root}")

    # Pre-resolve each show's destination folder.
    resolved_dirs: dict[str, Tuple[Path, bool]] = {}  # label -> (path, matched)
    for x in listed:
        matched = (
            find_show_folder(folder_idx, x.spec.label, tmdb_id=x.spec.tmdb_id)
            if folder_idx else None
        )
        if matched is not None:
            resolved_dirs[x.spec.label] = (matched, True)
        else:
            resolved_dirs[x.spec.label] = (args.out / safe_show_dir(x.spec.label), False)

    if args.save_mapping:
        with args.save_mapping.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["show", "matched", "folder"])
            for x in listed:
                folder, matched = resolved_dirs[x.spec.label]
                w.writerow([x.spec.label, "yes" if matched else "no", str(folder)])
        print(f"Wrote mapping CSV: {args.save_mapping}")

    if args.media_root:
        unmatched = [x.spec.label for x in listed
                     if not resolved_dirs[x.spec.label][1] and x.picked]
        if unmatched:
            print(f"WARN: {len(unmatched)} show(s) had no folder match "
                  f"under --media-root; falling back to {args.out}/:")
            for label in unmatched[:10]:
                print(f"  ? {label}")
            if len(unmatched) > 10:
                print(f"  ... and {len(unmatched) - 10} more "
                      f"(see --save-mapping CSV for the full list)")

    if args.dry_run:
        for x in listed:
            folder, matched = resolved_dirs[x.spec.label]
            tag = "->" if matched else "??"
            print(f"{x.spec.label}: {len(x.picked)} listed  {tag} {folder}")
        return

    dl_session = make_session(api_key)
    total_done = 0
    stop = False

    # Robust S/E parser: matches Show.S01E01.HDTV.x264.srt and friends, not
    # just the canonical S01E01.srt naming. Used to detect existing SRTs
    # regardless of who named them, so we don't burn quota re-downloading.
    _SE_IN_SRT_RE = re.compile(
        r"[Ss](\d{1,2})[\s._-]*[Ee](\d{1,3})|"
        r"(?<!\d)(\d{1,2})x(\d{1,3})(?!\d)"
    )

    def existing_se_in_folder(folder: Path) -> set:
        """Return the set of (season, episode) pairs already represented by
        ANY .srt file under `folder` (recursive), regardless of naming."""
        found = set()
        if not folder.is_dir():
            return found
        for p in folder.rglob("*.srt"):
            m = _SE_IN_SRT_RE.search(p.name)
            if m:
                if m.group(1):
                    found.add((int(m.group(1)), int(m.group(2))))
                else:
                    found.add((int(m.group(3)), int(m.group(4))))
        return found

    for show in listed:
        if stop:
            break
        if not show.picked:
            continue
        out_dir, _matched = resolved_dirs[show.spec.label]
        out_dir.mkdir(parents=True, exist_ok=True)

        existing_se = existing_se_in_folder(out_dir)

        if args.skip_complete:
            # Compare against count of unique (S, E) on disk, not file count
            if len(existing_se) >= len(show.picked):
                continue

        targets = show.picked
        if args.max_per_show > 0:
            targets = targets[: args.max_per_show]

        pending = []
        for sub in targets:
            attrs = sub.get("attributes", {})
            feat = attrs.get("feature_details", {}) or {}
            season = int(feat.get("season_number"))
            episode = int(feat.get("episode_number"))
            files = attrs.get("files", []) or []
            if not files:
                continue
            # Skip if we already have a sub for this (season, episode) under
            # ANY filename -- prevents creating canonical-named duplicates.
            if (season, episode) in existing_se:
                continue
            file_id = files[0].get("file_id")
            target = out_dir / f"S{season:02d}E{episode:02d}.srt"
            if target.exists():
                continue
            pending.append((target, file_id))

        if not pending:
            continue

        print(f"{show.spec.label}: {len(pending)} new subs to download")
        for target, file_id in tqdm(pending, desc=show.spec.label[:40],
                                    unit="sub", leave=False):
            if args.max_total_downloads and total_done >= args.max_total_downloads:
                print(f"\nHit --max-total-downloads "
                      f"({args.max_total_downloads}); stopping.")
                stop = True
                break
            try:
                data = download_one(dl_session, file_id)
            except QuotaExhausted as e:
                print(f"\nOpenSubtitles quota exhausted: {e}\n"
                      f"Stopping cleanly. Re-run tomorrow to resume.")
                stop = True
                break
            except Exception as e:
                tqdm.write(f"  ! download failed "
                           f"({show.spec.label} {target.name}): {e}")
                time.sleep(args.rate_delay)
                continue
            target.write_bytes(data)
            total_done += 1
            time.sleep(args.rate_delay)

    print(f"\nDone. Downloaded {total_done} new subtitle files.")
    print(f"Output root: {args.out}")


if __name__ == "__main__":
    main()
