#!/usr/bin/env python3
"""
Find show folders under <media-root> that aren't in your series_stats xlsx,
look them up on TMDB to get canonical title + year, and emit a CSV that
fetch_subtitles.py can consume.

Reads each folder's {tmdb-NNNN} marker to know which TMDB id to use, so
this only works for already-Plex-formatted folders.

Optionally also appends the rows to your xlsx (--append-xlsx) so future
batch runs of fetch_subtitles.py pick them up automatically. The append
mode preserves the existing sheet -- it adds rows at the bottom of the
summary sheet only and won't touch the detail sheet.

Usage:
    # Just emit a CSV
    python find_missing_shows.py "G:\\TV" --xlsx series_stats_detail.xlsx ^
        --out missing_shows.csv

    # CSV + append into xlsx summary sheet
    python find_missing_shows.py "G:\\TV" --xlsx series_stats_detail.xlsx ^
        --out missing_shows.csv --append-xlsx

Requires TMDB_API_KEY env var.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path

import requests
from tqdm import tqdm


TMDB_API = "https://api.themoviedb.org/3"
TMDB_FOLDER_RE = re.compile(r"\{tmdb-(\d+)\}", re.IGNORECASE)


def tmdb_show(api_key: str, tmdb_id: str) -> dict:
    """Return basic TMDB show details, or {} on failure."""
    try:
        r = requests.get(f"{TMDB_API}/tv/{tmdb_id}",
                         params={"api_key": api_key}, timeout=30)
        if r.status_code == 404:
            return {}
        r.raise_for_status()
        return r.json()
    except requests.RequestException:
        return {}


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("media_root", type=Path)
    ap.add_argument("--xlsx", type=Path, required=True,
                    help="series_stats_detail.xlsx")
    ap.add_argument("--summary-sheet", default="series_stats_summary")
    ap.add_argument("--out", type=Path, default=Path("missing_shows.csv"),
                    help="Where to write the fetch_subtitles-compatible CSV")
    ap.add_argument("--append-xlsx", action="store_true",
                    help="Also append the missing rows to the xlsx summary sheet")
    args = ap.parse_args()

    api_key = os.environ.get("TMDB_API_KEY")
    if not api_key:
        sys.exit("Set TMDB_API_KEY env var")

    if not args.media_root.is_dir():
        sys.exit(f"Not a directory: {args.media_root}")
    if not args.xlsx.is_file():
        sys.exit(f"xlsx not found: {args.xlsx}")

    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")

    # 1. Collect known TMDB ids from the xlsx
    wb = openpyxl.load_workbook(str(args.xlsx), data_only=True, read_only=True)
    ws = wb[args.summary_sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    try:
        id_col = header.index("tmdb_id")
    except ValueError:
        sys.exit(f"xlsx summary sheet has no 'tmdb_id' column. Header: {header}")
    known = set()
    for r in rows_iter:
        if r and r[id_col] is not None:
            known.add(str(r[id_col]))
    wb.close()
    print(f"xlsx has {len(known)} shows on the summary sheet.")

    # 2. Walk media_root for folders with {tmdb-NNNN} not in known
    missing = []
    for child in sorted(args.media_root.iterdir()):
        if not child.is_dir():
            continue
        m = TMDB_FOLDER_RE.search(child.name)
        if not m:
            continue
        tid = m.group(1)
        if tid in known:
            continue
        missing.append((tid, child.name))

    if not missing:
        print("No missing shows. Everything on disk is in the xlsx.")
        return
    print(f"Found {len(missing)} folder(s) under {args.media_root} not in xlsx.")

    # 3. TMDB lookup for each, build canonical rows
    rows_out = []
    for tid, folder_name in tqdm(missing, desc="tmdb lookup", unit="show"):
        meta = tmdb_show(api_key, tid)
        if not meta:
            print(f"  ! TMDB returned nothing for tmdb-{tid} ({folder_name})")
            rows_out.append({
                "tmdb_id": tid,
                "title": folder_name.split(" (")[0],   # best-effort
                "year": "",
                "status": "unknown",
                "folder": folder_name,
            })
            continue
        first_air = (meta.get("first_air_date") or "")[:4]
        rows_out.append({
            "tmdb_id": tid,
            "title": meta.get("name", folder_name),
            "year": first_air if first_air.isdigit() else "",
            "status": meta.get("status", ""),
            "folder": folder_name,
        })

    # 4. Write CSV that fetch_subtitles --from-csv understands
    fieldnames = ["tmdb_id", "title", "year", "status", "folder"]
    with args.out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows_out:
            w.writerow(r)
    print(f"Wrote {len(rows_out)} rows to {args.out}")

    # 5. Optionally append to xlsx
    if args.append_xlsx:
        wb = openpyxl.load_workbook(str(args.xlsx))
        ws = wb[args.summary_sheet]
        # Recompute header from the writable workbook
        full_header = [c.value for c in ws[1]]
        for r in rows_out:
            new_row = []
            for h in full_header:
                if h == "tmdb_id":
                    new_row.append(int(r["tmdb_id"]) if r["tmdb_id"].isdigit() else r["tmdb_id"])
                elif h == "title":
                    new_row.append(r["title"])
                elif h == "year":
                    new_row.append(int(r["year"]) if r["year"].isdigit() else r["year"] or None)
                elif h == "status":
                    new_row.append(r["status"] or None)
                else:
                    new_row.append(None)
            ws.append(new_row)
        wb.save(str(args.xlsx))
        print(f"Appended {len(rows_out)} rows to {args.xlsx} (sheet '{args.summary_sheet}').")

    # 6. Tell user the fetch command
    print()
    print(f"Next, fetch subtitles for these shows:")
    print(f"  python fetch_subtitles.py --from-csv {args.out} \\")
    print(f"      --title-col title --tmdb-col tmdb_id \\")
    print(f"      --media-root \"{args.media_root}\" --out subs --workers 6")


if __name__ == "__main__":
    main()
