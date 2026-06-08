#!/usr/bin/env python3
"""
Update series_stats_summary and series_stats_detail sheets IN PLACE in
your tracking xlsx, using current state on disk + TMDB. Preserves:

  - All other sheets (Summary, Doctor Who Checklist, my_movies, etc.)
  - Manual columns you've added (e.g. 'Complete  But Not Indexed?')
  - Formula cells (won't overwrite a cell whose value starts with '=')
  - All formatting / styles / column widths / cell merges

The script makes a timestamped .bak copy before modifying anything, so
you can always revert.

Updates are matched by tmdb_id (summary sheet) or (tmdb_id, season)
(detail sheet). Rows whose tmdb_id has no folder on disk are LEFT ALONE
-- those are wishlist / planning rows.

Requires TMDB_API_KEY env var.

Usage:
    python update_stats_xlsx.py "G:\\TV" series_stats_details.xlsx
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from tqdm import tqdm


TMDB_API = "https://api.themoviedb.org/3"
TMDB_FOLDER_RE = re.compile(r"\{tmdb-(\d+)\}", re.IGNORECASE)
SE_RE = re.compile(r"[Ss](\d{1,2})[\s._-]*[Ee](\d{1,3})")
SEASON_FOLDER_RE = re.compile(r"^[Ss]eason\s+(\d{1,2})(?:\s+\(\d{4}\))?\s*$")

VIDEO_EXTS = {".mkv", ".mp4", ".avi", ".mov", ".m4v", ".ts", ".m2ts",
              ".mpg", ".mpeg", ".wmv", ".webm"}

DURATION_CACHE_PATH = Path(__file__).parent / ".duration_cache.sqlite"
EXCEL_DURATION_FORMAT = "[h]:mm:ss"

# Columns we generate. Manual columns (anything NOT in this list) are
# preserved as-is, including a cell that contains a formula (starts with '=').
SUMMARY_WRITABLE = {
    "seasons_expected", "seasons_have",
    "regular_episodes_expected", "regular_episodes_have",
    "completeness_pct",
    "specials_expected", "specials_have",
    "missing_seasons", "incomplete_seasons", "unexpected_seasons",
    "status", "year",
    "total_runtime",
}
# Columns we'll auto-append to the summary header if the user hasn't added
# them yet, so first-time runs don't silently drop new data.
SUMMARY_AUTO_ADD = ["total_runtime"]

DETAIL_WRITABLE = {
    "season_year", "episodes_expected", "episodes_have",
    "missing_episodes", "extra_episodes", "status",
}


def open_duration_cache() -> sqlite3.Connection:
    """Open (and lazily create) the persistent ffprobe-duration cache.

    Keyed on (path, size, mtime) so files are re-probed only when they
    actually change on disk. Reuse across runs makes incremental updates
    nearly free.
    """
    conn = sqlite3.connect(str(DURATION_CACHE_PATH))
    conn.execute(
        "CREATE TABLE IF NOT EXISTS durations ("
        "  path  TEXT PRIMARY KEY,"
        "  size  INTEGER NOT NULL,"
        "  mtime REAL    NOT NULL,"
        "  seconds REAL  NOT NULL"
        ")"
    )
    conn.commit()
    return conn


_FFPROBE_WARNED = False


def safe_db_key(path: Path) -> str:
    """SQLite TEXT-safe stringification of a path.

    Windows tolerates lone Unicode surrogates in filenames (e.g. \\udd2e),
    but SQLite's default UTF-8 text adapter raises UnicodeEncodeError when
    binding such a string. We replace any unencodable code units with the
    Unicode replacement character so the cache still keys cleanly.
    """
    s = str(path)
    try:
        s.encode("utf-8")
        return s
    except UnicodeEncodeError:
        return s.encode("utf-8", "replace").decode("utf-8", "replace")


def probe_duration_seconds(path: Path, conn: sqlite3.Connection) -> float:
    """Return container duration in seconds. 0.0 if ffprobe fails or is missing.

    Caches successful probes by (safe_db_key(path), size, mtime). A file
    whose size+mtime are unchanged since the last successful probe is
    served from cache. Failed probes are NOT cached, so retries pick them
    up on the next run.
    """
    global _FFPROBE_WARNED
    try:
        st = path.stat()
    except OSError:
        return 0.0
    key = safe_db_key(path)
    try:
        row = conn.execute(
            "SELECT size, mtime, seconds FROM durations WHERE path = ?", (key,)
        ).fetchone()
    except sqlite3.Error:
        row = None
    if row and row[0] == st.st_size and abs(row[1] - st.st_mtime) < 1.0:
        return float(row[2])

    # NOTE: We intentionally run ffprobe in BYTES mode (no text=True/encoding=).
    # The Windows default cp1252 codec cannot decode UTF-8 filenames that
    # ffprobe echoes to stderr, and that crashes the reader thread, losing
    # stdout entirely. Decoding manually with errors="replace" is bulletproof.
    seconds: Optional[float] = None
    try:
        result = subprocess.run(
            [
                "ffprobe", "-v", "error",
                "-show_entries", "format=duration",
                "-of", "default=noprint_wrappers=1:nokey=1",
                str(path),
            ],
            capture_output=True, timeout=30,
        )
        out = result.stdout.decode("utf-8", errors="replace").strip()
        if out and out.lower() != "n/a":
            try:
                seconds = float(out)
            except ValueError:
                seconds = None
    except FileNotFoundError:
        if not _FFPROBE_WARNED:
            print("WARN: ffprobe not found on PATH; runtimes will be 0",
                  file=sys.stderr)
            _FFPROBE_WARNED = True
        return 0.0
    except subprocess.TimeoutExpired:
        seconds = None
    except OSError:
        seconds = None

    if seconds is None:
        # Don't cache failures -- next run will retry rather than memoize 0.
        return 0.0

    try:
        conn.execute(
            "INSERT OR REPLACE INTO durations (path, size, mtime, seconds) "
            "VALUES (?, ?, ?, ?)",
            (key, st.st_size, st.st_mtime, seconds),
        )
    except sqlite3.Error:
        # Cache write failed (e.g. exotic key collision). Probe value is
        # still returned to the caller, so this run isn't affected.
        pass
    return seconds


def purge_zero_cached(conn: sqlite3.Connection) -> int:
    """Drop cached seconds=0 rows so they get re-probed.

    Useful after the cp1252 reader-thread bug poisoned the cache with
    bogus 0-second entries. Returns rowcount deleted.
    """
    cur = conn.execute("DELETE FROM durations WHERE seconds = 0")
    conn.commit()
    return cur.rowcount or 0


def seconds_to_excel_duration(seconds: float) -> float:
    """Excel stores durations as a fraction of a day (86400 s)."""
    if not seconds or seconds < 0:
        return 0.0
    return float(seconds) / 86400.0


def scan_folder_video_stats(
    folder: Path, conn: sqlite3.Connection,
) -> Tuple[int, float]:
    """Walk folder; return (video_file_count, total_seconds)."""
    count = 0
    total = 0.0
    for p in folder.rglob("*"):
        if not p.is_file() or p.suffix.lower() not in VIDEO_EXTS:
            continue
        count += 1
        total += probe_duration_seconds(p, conn)
    return count, total


def tmdb_show(sess: requests.Session, tmdb_id: str) -> Optional[dict]:
    try:
        r = sess.get(f"{TMDB_API}/tv/{tmdb_id}", timeout=30)
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.json()
    except requests.RequestException:
        return None


def scan_show_disk(
    show_dir: Path, conn: sqlite3.Connection,
) -> Tuple[Dict[int, set], int, float]:
    """Walk show_dir once; return (found_by_season, video_count, total_seconds).

    found_by_season is {season: {episode, ...}} parsed from filenames /
    parent season folders. video_count and total_seconds cover ALL video
    files under show_dir (including extras), which is what we want for
    the new total_runtime column.
    """
    found: Dict[int, set] = defaultdict(set)
    video_count = 0
    total_seconds = 0.0
    for p in show_dir.rglob("*"):
        if not p.is_file() or p.suffix.lower() not in VIDEO_EXTS:
            continue
        video_count += 1
        total_seconds += probe_duration_seconds(p, conn)
        m = SE_RE.search(p.name)
        if m:
            found[int(m.group(1))].add(int(m.group(2)))
            continue
        mseason = SEASON_FOLDER_RE.match(p.parent.name)
        if mseason:
            s = int(mseason.group(1))
            mep = re.search(r"(?:^|[^0-9])(\d{1,3})(?:[^0-9]|$)", p.name)
            if mep:
                found[s].add(int(mep.group(1)))
    return found, video_count, total_seconds


def compact_range(missing: List[int]) -> str:
    if not missing:
        return ""
    runs: List[Tuple[int, int]] = []
    start = prev = missing[0]
    for n in missing[1:]:
        if n == prev + 1:
            prev = n
        else:
            runs.append((start, prev))
            start = prev = n
    runs.append((start, prev))
    return ", ".join(f"{a}" if a == b else f"{a}-{b}" for a, b in runs)


def compute_show_stats(
    meta: dict,
    have: Dict[int, set],
    video_count: int = 0,
    total_seconds: float = 0.0,
) -> dict:
    """Build the summary row dict for one show from TMDB metadata + disk scan.

    video_count and total_seconds cover ALL video files under the show
    directory -- episodes, specials, extras, behind-the-scenes, etc. --
    which is what the new total_runtime column reports.
    """
    expected: Dict[int, Tuple[int, Optional[int]]] = {}
    for s in meta.get("seasons") or []:
        snum = s.get("season_number")
        if snum is None:
            continue
        ec = s.get("episode_count") or 0
        air = (s.get("air_date") or "")[:4]
        sy = int(air) if air.isdigit() else None
        expected[int(snum)] = (int(ec), sy)

    regular = sorted(s for s in expected if s > 0)
    seasons_expected = len(regular)
    seasons_have = sum(1 for s in regular if have.get(s))
    reg_exp = sum(expected[s][0] for s in regular)
    reg_have = sum(len(have.get(s, set())) for s in regular)
    spec_exp = expected.get(0, (0, None))[0]
    spec_have = len(have.get(0, set()))
    pct = round(reg_have / reg_exp * 100, 1) if reg_exp else 0.0

    missing_seasons = []
    incomplete_seasons = []
    unexpected_seasons = []
    for s in regular:
        exp, _ = expected[s]
        h = len(have.get(s, set()))
        if h == 0 and exp > 0:
            missing_seasons.append(s)
        elif h < exp:
            incomplete_seasons.append(f"S{s:02d}({h}/{exp})")
    for s in have:
        if s not in expected:
            unexpected_seasons.append(s)

    first_air = (meta.get("first_air_date") or "")[:4]
    year = int(first_air) if first_air.isdigit() else None
    return {
        "year": year,
        "status": meta.get("status", ""),
        "seasons_expected": seasons_expected,
        "seasons_have": seasons_have,
        "regular_episodes_expected": reg_exp,
        "regular_episodes_have": reg_have,
        "completeness_pct": pct,
        "specials_expected": spec_exp,
        "specials_have": spec_have,
        "missing_seasons": ", ".join(str(s) for s in missing_seasons),
        "incomplete_seasons": ", ".join(incomplete_seasons),
        "unexpected_seasons": ", ".join(str(s) for s in sorted(unexpected_seasons)),
        "total_runtime": float(total_seconds),  # raw seconds; rendered by caller
        "_expected_by_season": expected,
        "_have_by_season": have,
        "_video_count": video_count,
        "_total_seconds": float(total_seconds),
    }


def write_cell_safe(
    ws, row: int, col: int, value, number_format: Optional[str] = None,
) -> bool:
    """Write to cell unless it currently contains a formula (string starting
    with '='). Returns True if written. Applies number_format when given."""
    cell = ws.cell(row=row, column=col)
    existing = cell.value
    if isinstance(existing, str) and existing.startswith("="):
        return False
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return True


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("media_root", type=Path)
    ap.add_argument("xlsx", type=Path,
                    help="Path to the tracking xlsx to update in place")
    ap.add_argument("--summary-sheet", default="series_stats_summary")
    ap.add_argument("--detail-sheet", default="series_stats_detail")
    ap.add_argument("--unmapped-sheet", default="Unmapped TV Series",
                    help="Sheet name for root folders without a {tmdb-#####} tag")
    ap.add_argument("--no-backup", action="store_true",
                    help="Skip the .bak copy before modifying (NOT recommended)")
    ap.add_argument("--skip-unmapped", action="store_true",
                    help="Skip generating the Unmapped TV Series sheet")
    ap.add_argument("--autosave-every", type=int, default=25,
                    help="Save the workbook every N processed shows "
                         "(default 25; set to 0 to disable)")
    ap.add_argument("--reprobe-zeros", action="store_true",
                    help="Delete cached 0-second rows before running so they "
                         "get re-probed (use after the cp1252 reader-thread "
                         "bug poisoned the cache).")
    args = ap.parse_args()

    api_key = os.environ.get("TMDB_API_KEY")
    if not api_key:
        sys.exit("Set TMDB_API_KEY env var")

    if not args.media_root.is_dir():
        sys.exit(f"Not a directory: {args.media_root}")
    if not args.xlsx.is_file():
        sys.exit(f"xlsx not found: {args.xlsx}")

    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")

    # 1. Backup
    if not args.no_backup:
        ts = time.strftime("%Y%m%d-%H%M%S")
        bak = args.xlsx.with_suffix(f".bak-{ts}.xlsx")
        shutil.copy2(args.xlsx, bak)
        print(f"Backup written: {bak}")

    # 2. Build TMDB-id -> show folder index, AND collect unmapped folders
    folder_for_tmdb: Dict[str, Path] = {}
    unmapped_folders: List[Path] = []
    for child in args.media_root.iterdir():
        if not child.is_dir():
            continue
        m = TMDB_FOLDER_RE.search(child.name)
        if m:
            folder_for_tmdb[m.group(1)] = child
        else:
            unmapped_folders.append(child)
    print(f"Show folders under {args.media_root}: "
          f"{len(folder_for_tmdb)} mapped, {len(unmapped_folders)} unmapped")

    # 3. Load xlsx (NOT data_only, so we can see formulas to preserve them)
    wb = openpyxl.load_workbook(str(args.xlsx))
    sess = requests.Session()
    sess.params = {"api_key": api_key}

    # Persistent ffprobe-duration cache (keyed on path+size+mtime).
    dur_conn = open_duration_cache()

    if args.reprobe_zeros:
        purged = purge_zero_cached(dur_conn)
        print(f"Purged {purged} zero-second cache rows; they'll be re-probed.")

    show_stats_cache: Dict[str, dict] = {}

    # 4. Update summary sheet
    if args.summary_sheet not in wb.sheetnames:
        print(f"WARN: no sheet named {args.summary_sheet!r}; skipping summary")
    else:
        ws = wb[args.summary_sheet]
        header = [c.value for c in ws[1]]
        col_idx = {h: i + 1 for i, h in enumerate(header) if h}
        if "tmdb_id" not in col_idx:
            sys.exit(f"{args.summary_sheet} has no tmdb_id column")

        # Auto-append any new columns we generate that the user hasn't
        # already added (preserves their existing column layout).
        next_col = (ws.max_column or len(header)) + 1
        for new_col in SUMMARY_AUTO_ADD:
            if new_col not in col_idx:
                ws.cell(row=1, column=next_col, value=new_col)
                col_idx[new_col] = next_col
                print(f"  appended new column {new_col!r} at column {next_col}")
                next_col += 1

        updated = 0
        skipped_no_folder = 0
        skipped_no_tmdb = 0
        formula_preserved = 0
        autosave_every = max(0, int(args.autosave_every or 0))
        processed_since_save = 0
        for r in tqdm(range(2, ws.max_row + 1), desc=args.summary_sheet,
                      unit="row"):
            tid = ws.cell(row=r, column=col_idx["tmdb_id"]).value
            if not tid:
                continue
            tid = str(int(tid)) if isinstance(tid, (int, float)) else str(tid).strip()
            folder = folder_for_tmdb.get(tid)
            if folder is None:
                skipped_no_folder += 1
                continue
            if tid not in show_stats_cache:
                meta = tmdb_show(sess, tid)
                if not meta:
                    show_stats_cache[tid] = {}
                else:
                    have, vcount, tsec = scan_show_disk(folder, dur_conn)
                    show_stats_cache[tid] = compute_show_stats(
                        meta, have, vcount, tsec,
                    )
            stats = show_stats_cache[tid]
            if not stats:
                skipped_no_tmdb += 1
                continue
            wrote_anything = False
            for col_name, value in stats.items():
                if col_name.startswith("_"):
                    continue  # internal helpers
                if col_name not in SUMMARY_WRITABLE:
                    continue
                if col_name not in col_idx:
                    continue
                if col_name == "total_runtime":
                    # Render as an Excel duration so [h]:mm:ss sorts numerically.
                    excel_val = seconds_to_excel_duration(value)
                    wrote = write_cell_safe(
                        ws, r, col_idx[col_name], excel_val,
                        number_format=EXCEL_DURATION_FORMAT,
                    )
                else:
                    wrote = write_cell_safe(ws, r, col_idx[col_name], value)
                if wrote:
                    wrote_anything = True
                else:
                    formula_preserved += 1
            if wrote_anything:
                updated += 1
                processed_since_save += 1
                if autosave_every and processed_since_save >= autosave_every:
                    try:
                        dur_conn.commit()
                        wb.save(str(args.xlsx))
                        processed_since_save = 0
                    except PermissionError as exc:
                        tqdm.write(
                            f"WARN: autosave to {args.xlsx} failed "
                            f"(is the file open in Excel?): {exc}"
                        )

        # Final commit of cache + workbook for this section.
        try:
            dur_conn.commit()
            wb.save(str(args.xlsx))
        except PermissionError as exc:
            print(f"WARN: end-of-summary save failed: {exc}", file=sys.stderr)

        print(f"  {args.summary_sheet}: updated={updated}, "
              f"skipped(no folder on disk)={skipped_no_folder}, "
              f"skipped(no tmdb data)={skipped_no_tmdb}, "
              f"formula cells preserved={formula_preserved}")

    # 5. Update detail sheet
    if args.detail_sheet not in wb.sheetnames:
        print(f"WARN: no sheet named {args.detail_sheet!r}; skipping detail")
    else:
        ws = wb[args.detail_sheet]
        header = [c.value for c in ws[1]]
        col_idx = {h: i + 1 for i, h in enumerate(header) if h}
        for required in ("tmdb_id", "season"):
            if required not in col_idx:
                sys.exit(f"{args.detail_sheet} has no {required} column")

        updated = 0
        skipped_no_folder = 0
        skipped_no_tmdb = 0
        formula_preserved = 0
        for r in tqdm(range(2, ws.max_row + 1), desc=args.detail_sheet,
                      unit="row"):
            tid = ws.cell(row=r, column=col_idx["tmdb_id"]).value
            s_val = ws.cell(row=r, column=col_idx["season"]).value
            if not tid or s_val is None:
                continue
            tid = str(int(tid)) if isinstance(tid, (int, float)) else str(tid).strip()
            try:
                season = int(s_val)
            except (ValueError, TypeError):
                continue
            folder = folder_for_tmdb.get(tid)
            if folder is None:
                skipped_no_folder += 1
                continue
            if tid not in show_stats_cache:
                meta = tmdb_show(sess, tid)
                if not meta:
                    show_stats_cache[tid] = {}
                else:
                    have, vcount, tsec = scan_show_disk(folder, dur_conn)
                    show_stats_cache[tid] = compute_show_stats(
                        meta, have, vcount, tsec,
                    )
            stats = show_stats_cache[tid]
            if not stats:
                skipped_no_tmdb += 1
                continue
            expected = stats["_expected_by_season"].get(season, (0, None))
            have_set = stats["_have_by_season"].get(season, set())
            exp_count, season_year = expected
            have_count = len(have_set)
            if exp_count == 0 and have_count == 0:
                row_status = "no_data"
            elif exp_count == 0 and have_count > 0:
                row_status = "unexpected"
            elif have_count == 0:
                row_status = "missing"
            elif have_count < exp_count:
                row_status = "incomplete"
            elif have_count == exp_count:
                row_status = "complete"
            else:
                row_status = "extra"
            missing_eps = compact_range(sorted(
                set(range(1, exp_count + 1)) - have_set
            )) if exp_count else ""
            extra_eps = compact_range(sorted(
                have_set - set(range(1, exp_count + 1))
            )) if exp_count else (compact_range(sorted(have_set)) if have_set else "")
            updates = {
                "season_year": season_year if season_year else None,
                "episodes_expected": exp_count,
                "episodes_have": have_count,
                "missing_episodes": missing_eps,
                "extra_episodes": extra_eps,
                "status": row_status,
            }
            wrote_anything = False
            for col_name, value in updates.items():
                if col_name not in DETAIL_WRITABLE:
                    continue
                if col_name not in col_idx:
                    continue
                if write_cell_safe(ws, r, col_idx[col_name], value):
                    wrote_anything = True
                else:
                    formula_preserved += 1
            if wrote_anything:
                updated += 1

        print(f"  {args.detail_sheet}: updated={updated}, "
              f"skipped(no folder on disk)={skipped_no_folder}, "
              f"skipped(no tmdb data)={skipped_no_tmdb}, "
              f"formula cells preserved={formula_preserved}")

    # 6. Unmapped TV Series sheet
    if args.skip_unmapped:
        print(f"Skipping unmapped-folder scan ({args.unmapped_sheet!r}).")
    elif not unmapped_folders:
        print(f"No unmapped root folders found; "
              f"{args.unmapped_sheet!r} sheet not written.")
    else:
        sheet_name = args.unmapped_sheet
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_u = wb.create_sheet(sheet_name)
        ws_u.append(["directory", "video_files", "total_runtime"])
        for cell in ws_u[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        unmapped_total_videos = 0
        unmapped_total_seconds = 0.0
        for folder in tqdm(
            sorted(unmapped_folders, key=lambda p: p.name.lower()),
            desc=sheet_name, unit="dir",
        ):
            count, seconds = scan_folder_video_stats(folder, dur_conn)
            unmapped_total_videos += count
            unmapped_total_seconds += seconds
            dur_cell_val = seconds_to_excel_duration(seconds)
            ws_u.append([folder.name, count, dur_cell_val])
            ws_u.cell(
                row=ws_u.max_row, column=3,
            ).number_format = EXCEL_DURATION_FORMAT
        # Reasonable column widths so the user doesn't have to autosize by hand.
        ws_u.column_dimensions["A"].width = 60
        ws_u.column_dimensions["B"].width = 14
        ws_u.column_dimensions["C"].width = 16
        ws_u.freeze_panes = "A2"
        print(f"  {sheet_name}: {len(unmapped_folders)} folders, "
              f"{unmapped_total_videos} video files, "
              f"{unmapped_total_seconds / 3600:.1f} hours total")

    # 7. Close the duration cache and save the workbook
    dur_conn.commit()
    dur_conn.close()
    wb.save(str(args.xlsx))
    print(f"\nSaved: {args.xlsx}")


if __name__ == "__main__":
    main()
