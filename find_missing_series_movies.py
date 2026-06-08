#!/usr/bin/env python3
r"""
find_missing_series_movies.py
=============================

Walk a local movie library, find folders that group movies into a series,
look every movie up on TMDB by its IMDb ID, and report which films of each
TMDB "collection" are missing locally.

Expected directory layout (your layout):

    H:\Movies\<SeriesName>\<Movie Name> <year> imdb-tt########\<file>
    H:\Movies\<Movie Name> <year> imdb-tt########\<file>            (skipped - not in a series)

A folder is treated as a "series folder" when it is NOT itself a movie folder
(no `<year> imdb-tt########`) AND it contains at least one subfolder that IS
a movie folder.

Usage
-----
    set TMDB_API_KEY=your_v3_key_here
    python find_missing_series_movies.py H:\Movies H:\_mediaprep\series_stats_details.xlsx

The second argument may be either:
  * a `.csv` path - fresh CSV is written each run, or
  * a `.xlsx` path - the script will MERGE into an existing workbook,
    preserving any notes/highlights you've added in extra columns.

Optional flags:
    --sheet <name>           Sheet to update in the xlsx (default: "missing movies").
    --include-unreleased     Include collection entries whose release_date is
                             in the future or missing (default: skip them).
    --cache <path>           Cache file (default: %USERPROFILE%\.tmdb_series_cache.json).
    --sleep <seconds>        Sleep between TMDB calls (default: 0.05).

Get a free TMDB v3 API key at https://www.themoviedb.org/settings/api
(create an account, request a developer key; it's instant).

Output
------
CSV mode: one row per missing movie with columns:
    collection_name, series_folder(s), missing_title, missing_year,
    missing_imdb_id, missing_tmdb_id, release_date

XLSX mode (merge):
- Existing rows whose movie is still missing locally are left untouched
  (so any notes/highlights in extra columns are preserved).
- Existing rows whose movie now exists in your library are DELETED.
- New missing movies discovered on this run are APPENDED.
- Rows are keyed by `missing_imdb_id` (falling back to `missing_tmdb_id`).
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    sys.stderr.write(
        "This script needs the 'requests' package. Install it with:\n"
        "    pip install requests\n"
    )
    sys.exit(2)


TMDB_BASE = "https://api.themoviedb.org/3"
IMDB_RE = re.compile(r"imdb-(tt\d{7,10})", re.IGNORECASE)
YEAR_RE = re.compile(r"(?<!\d)(19|20)\d{2}(?!\d)")


# ---------------------------------------------------------------------------
# Filesystem walking
# ---------------------------------------------------------------------------

def looks_like_movie_folder(name: str) -> bool:
    """A movie folder has both a 4-digit year and an imdb-tt id in its name."""
    return bool(IMDB_RE.search(name) and YEAR_RE.search(name))


def extract_imdb_id(name: str) -> str | None:
    m = IMDB_RE.search(name)
    return m.group(1).lower() if m else None


def walk_series(root: Path):
    """
    Yield (series_folder_name, [(imdb_id, movie_folder_name), ...]) for each
    top-level folder under `root` that contains movie subfolders.
    Top-level folders that look like movie folders themselves are skipped here
    (they're handled separately by walk_top_level_movies).
    """
    if not root.is_dir():
        raise SystemExit(f"Not a directory: {root}")

    for entry in sorted(root.iterdir(), key=lambda p: p.name.lower()):
        if not entry.is_dir():
            continue
        if looks_like_movie_folder(entry.name):
            continue

        movies = []
        try:
            children = list(entry.iterdir())
        except PermissionError:
            sys.stderr.write(f"warn: permission denied reading {entry}\n")
            continue

        for sub in sorted(children, key=lambda p: p.name.lower()):
            if sub.is_dir() and looks_like_movie_folder(sub.name):
                tt = extract_imdb_id(sub.name)
                if tt:
                    movies.append((tt, sub.name))

        if movies:
            yield entry.name, movies


def walk_top_level_movies(root: Path):
    """Yield (imdb_id, folder_name) for movies filed directly at the root.

    These don't define any new tracked collection on their own - but if one of
    them happens to belong to a collection already tracked via a series folder,
    we want to credit it so the missing-row gets cleared.
    """
    if not root.is_dir():
        return
    for entry in sorted(root.iterdir(), key=lambda p: p.name.lower()):
        if entry.is_dir() and looks_like_movie_folder(entry.name):
            tt = extract_imdb_id(entry.name)
            if tt:
                yield tt, entry.name


# ---------------------------------------------------------------------------
# TMDB client (cached)
# ---------------------------------------------------------------------------

class TMDB:
    def __init__(self, api_key: str, cache_path: Path, sleep_s: float = 0.05):
        self.api_key = api_key
        self.cache_path = cache_path
        self.sleep_s = sleep_s
        self.session = requests.Session()
        self.cache: dict = {}
        if cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text(encoding="utf-8"))
            except Exception:
                self.cache = {}

    def save(self) -> None:
        try:
            self.cache_path.write_text(json.dumps(self.cache), encoding="utf-8")
        except Exception as e:
            sys.stderr.write(f"warn: failed to save cache: {e}\n")

    def get(self, path: str, params: dict | None = None):
        params = dict(params or {})
        key = path + "?" + json.dumps(params, sort_keys=True)
        if key in self.cache:
            return self.cache[key]
        params["api_key"] = self.api_key
        url = TMDB_BASE + path
        for attempt in range(5):
            try:
                r = self.session.get(url, params=params, timeout=30)
            except requests.RequestException as e:
                if attempt == 4:
                    raise
                time.sleep(2 ** attempt)
                continue
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", "2")) + 1
                time.sleep(wait)
                continue
            if r.status_code == 404:
                # Do NOT cache negative responses persistently. New releases
                # sometimes 404 on /find/ until TMDB indexes them; we don't
                # want a one-time 404 to stick around forever.
                if self.sleep_s:
                    time.sleep(self.sleep_s)
                return None
            r.raise_for_status()
            data = r.json()
            self.cache[key] = data
            if self.sleep_s:
                time.sleep(self.sleep_s)
            return data
        return None

    # ----- semantic helpers -----

    def find_by_imdb(self, imdb_id: str):
        return self.get(f"/find/{imdb_id}", {"external_source": "imdb_id"})

    def movie(self, movie_id: int):
        return self.get(f"/movie/{movie_id}", {})

    def movie_external_ids(self, movie_id: int):
        return self.get(f"/movie/{movie_id}/external_ids", {})

    def collection(self, collection_id: int):
        return self.get(f"/collection/{collection_id}", {})


CSV_FIELDS = [
    "collection_name", "series_folder(s)",
    "missing_title", "missing_year",
    "missing_imdb_id", "missing_tmdb_id",
    "release_date",
]


# ---------------------------------------------------------------------------
# XLSX merge (preserves existing notes/highlights)
# ---------------------------------------------------------------------------

def _row_key(imdb_id, tmdb_id):
    """Stable key for matching rows across runs.

    Prefer IMDb id (cheap to read off a folder); fall back to TMDB id.
    Returns None if neither is usable.
    """
    if imdb_id:
        s = str(imdb_id).strip().lower()
        if s.startswith("tt"):
            return ("imdb", s)
    if tmdb_id not in (None, ""):
        try:
            return ("tmdb", int(tmdb_id))
        except (TypeError, ValueError):
            return ("tmdb", str(tmdb_id).strip())
    return None


def update_xlsx(xlsx_path: Path, sheet_name: str, new_rows: list[dict]):
    """Merge `new_rows` into `xlsx_path` / `sheet_name`.

    Returns a dict with counts: {"deleted", "added", "kept", "total_after"}.

    Behaviour:
    - Existing data rows are identified by (imdb_id, fallback tmdb_id).
    - Rows whose key is no longer in `new_rows` are DELETED (the movie
      moved into the library).
    - Rows whose key is still in `new_rows` are KEPT untouched - this
      preserves notes/highlights in any extra columns the user added.
    - Keys in `new_rows` not present in the sheet are APPENDED.
    - Missing standard headers are added at the end of row 1 without
      disturbing any extra columns of the user's own.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise SystemExit(
            "This script needs 'openpyxl' to update .xlsx files.\n"
            "Install it with:  pip install openpyxl"
        )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # New workbooks get a default "Sheet" we don't want.
        default = wb.active
        if default is not None:
            wb.remove(default)

    # Locate target sheet (case-insensitive); create if absent.
    ws = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.create_sheet(sheet_name)
        for c, h in enumerate(CSV_FIELDS, start=1):
            ws.cell(row=1, column=c, value=h)

    # Build header -> column index map (case-insensitive).
    header_to_col: dict[str, int] = {}
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            header_to_col[str(v).strip().lower()] = c

    # Make sure all expected headers exist; add any missing ones to the right
    # of whatever the user has there. We do not reorder existing columns.
    for h in CSV_FIELDS:
        if h.lower() not in header_to_col:
            new_col = (max(header_to_col.values()) if header_to_col else 0) + 1
            ws.cell(row=1, column=new_col, value=h)
            header_to_col[h.lower()] = new_col

    imdb_col = header_to_col["missing_imdb_id"]
    tmdb_col = header_to_col["missing_tmdb_id"]

    # Walk existing rows; collect keys -> row index.
    last_row = ws.max_row or 1
    existing: dict = {}
    for r in range(2, last_row + 1):
        imdb_val = ws.cell(row=r, column=imdb_col).value
        tmdb_val = ws.cell(row=r, column=tmdb_col).value
        # Treat completely empty rows as non-existent
        if (imdb_val in (None, "") and tmdb_val in (None, "")
                and all(ws.cell(row=r, column=c).value in (None, "")
                        for c in range(1, ws.max_column + 1))):
            continue
        key = _row_key(imdb_val, tmdb_val)
        if key is not None:
            # If duplicates exist, keep the FIRST (lowest row index).
            existing.setdefault(key, r)

    # Index incoming rows by key.
    new_by_key: dict = {}
    for r in new_rows:
        k = _row_key(r.get("missing_imdb_id"), r.get("missing_tmdb_id"))
        if k is not None:
            new_by_key.setdefault(k, r)

    existing_keys = set(existing.keys())
    new_keys = set(new_by_key.keys())

    to_delete_keys = existing_keys - new_keys
    to_add_keys = new_keys - existing_keys
    kept_keys = existing_keys & new_keys

    # Delete bottom-up so row indices above stay stable.
    delete_row_indices = sorted((existing[k] for k in to_delete_keys), reverse=True)
    for r in delete_row_indices:
        ws.delete_rows(r, 1)

    # Append new rows. Recompute the next free row after deletions.
    next_row = (ws.max_row or 1) + 1
    if next_row < 2:
        next_row = 2
    for k in sorted(to_add_keys, key=lambda x: (str(x[0]), str(x[1]))):
        row = new_by_key[k]
        for h in CSV_FIELDS:
            col = header_to_col[h.lower()]
            ws.cell(row=next_row, column=col, value=row.get(h))
        next_row += 1

    wb.save(xlsx_path)

    return {
        "deleted": len(to_delete_keys),
        "added": len(to_add_keys),
        "kept": len(kept_keys),
        "total_after": len(kept_keys) + len(to_add_keys),
    }


def resolve_collection(tmdb: TMDB, imdb_id: str):
    """Return (tmdb_movie_id, collection_id, collection_name) or None."""
    find = tmdb.find_by_imdb(imdb_id)
    if not find:
        return None
    results = find.get("movie_results") or []
    if not results:
        return None
    movie_id = results[0]["id"]
    movie = tmdb.movie(movie_id)
    if not movie:
        return None
    coll = movie.get("belongs_to_collection")
    if not coll:
        return None
    return movie_id, coll["id"], coll["name"]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("movies_root", help=r"Root of your library, e.g. H:\Movies")
    p.add_argument("out_path", help="Output path - .csv (fresh write) or .xlsx (merge)")
    p.add_argument("--sheet", default="missing movies",
                   help='Sheet name to update when output is .xlsx (default: "missing movies")')
    p.add_argument("--include-unreleased", action="store_true",
                   help="Include collection entries with no release date or a future release date")
    default_cache = Path(os.environ.get("USERPROFILE") or Path.home()) / ".tmdb_series_cache.json"
    p.add_argument("--cache", default=str(default_cache),
                   help=f"TMDB response cache file (default: {default_cache})")
    p.add_argument("--sleep", type=float, default=0.05,
                   help="Sleep between TMDB calls in seconds (default: 0.05)")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)

    api_key = os.environ.get("TMDB_API_KEY")
    if not api_key:
        sys.stderr.write(
            "error: TMDB_API_KEY environment variable is not set.\n"
            "       Get a free v3 key at https://www.themoviedb.org/settings/api\n"
            "       Then: set TMDB_API_KEY=your_key_here\n"
        )
        return 2

    movies_root = Path(args.movies_root)
    tmdb = TMDB(api_key, Path(args.cache), sleep_s=args.sleep)

    # collection_id -> dict
    collections: dict[int, dict] = {}
    unmatched_series: list[tuple[str, int]] = []  # (folder, n_movies)
    scanned_series_count = 0

    # ---- Pass 1: walk series folders, build the set of tracked collections.
    try:
        for series_folder, movies in walk_series(movies_root):
            scanned_series_count += 1
            print(f"[scan] {series_folder}  ({len(movies)} movies)")
            matched = 0
            for imdb_id, _movie_folder in movies:
                try:
                    info = resolve_collection(tmdb, imdb_id)
                except requests.HTTPError as e:
                    sys.stderr.write(f"  warn: TMDB error for {imdb_id}: {e}\n")
                    continue
                if not info:
                    continue
                movie_id, coll_id, coll_name = info
                matched += 1
                bucket = collections.setdefault(coll_id, {
                    "name": coll_name,
                    "have_tmdb": set(),
                    "have_imdb": set(),
                    "series_folders": set(),
                })
                bucket["have_tmdb"].add(movie_id)
                bucket["have_imdb"].add(imdb_id)
                bucket["series_folders"].add(series_folder)
            if matched == 0:
                unmatched_series.append((series_folder, len(movies)))
    finally:
        tmdb.save()

    # ---- Pass 2: for each tracked collection, fetch parts + external IDs.
    # We build imdb_to_tracked[imdb_id] -> (coll_id, tmdb_id), then in Pass 3
    # we can credit any top-level movie whose IMDb shows up in that map
    # without needing a TMDB call per top-level movie.
    parts_by_coll: dict[int, list[dict]] = {}
    imdb_to_tracked: dict[str, tuple[int, int]] = {}
    for coll_id, info in collections.items():
        try:
            coll_data = tmdb.collection(coll_id)
        except requests.HTTPError as e:
            sys.stderr.write(f"warn: failed to fetch collection {coll_id}: {e}\n")
            continue
        if not coll_data:
            continue
        parts = coll_data.get("parts") or []
        enriched = []
        for p in parts:
            try:
                ext = tmdb.movie_external_ids(p["id"]) or {}
            except requests.HTTPError:
                ext = {}
            imdb_id = (ext.get("imdb_id") or "").lower()
            enriched.append({
                "id": p["id"],
                "title": p.get("title") or p.get("original_title") or "",
                "release_date": p.get("release_date") or "",
                "imdb_id": imdb_id,
            })
            if imdb_id:
                imdb_to_tracked[imdb_id] = (coll_id, p["id"])
        parts_by_coll[coll_id] = enriched
        tmdb.save()

    # ---- Pass 3: walk top-level standalone movies, credit them against any
    # already-tracked collection. We deliberately do NOT introduce new tracked
    # collections from top-level movies - the user's series folders define
    # what counts as a tracked franchise.
    top_level_credit = 0
    for imdb_id, _folder in walk_top_level_movies(movies_root):
        tracked = imdb_to_tracked.get(imdb_id)
        if tracked is None:
            continue
        coll_id, tmdb_id = tracked
        info = collections[coll_id]
        if tmdb_id not in info["have_tmdb"]:
            info["have_tmdb"].add(tmdb_id)
            info["have_imdb"].add(imdb_id)
            top_level_credit += 1
    if top_level_credit:
        print(f"[top-level] credited {top_level_credit} standalone movie(s) "
              f"against tracked collections")

    # ---- Pass 4: diff. parts list per collection already in parts_by_coll.
    today = dt.date.today().isoformat()
    rows = []
    for coll_id, info in collections.items():
        for p in parts_by_coll.get(coll_id, []):
            if p["id"] in info["have_tmdb"]:
                continue
            release_date = p.get("release_date") or ""
            if not args.include_unreleased:
                if not release_date or release_date > today:
                    continue
            year = release_date[:4] if release_date else ""
            rows.append({
                "collection_name": info["name"],
                "series_folder(s)": " | ".join(sorted(info["series_folders"])),
                "missing_title": p.get("title") or "",
                "missing_year": year,
                "missing_imdb_id": p.get("imdb_id") or "",
                "missing_tmdb_id": p["id"],
                "release_date": release_date,
            })
    tmdb.save()

    rows.sort(key=lambda r: (r["collection_name"].lower(), r["missing_year"], r["missing_title"].lower()))

    out_path = Path(args.out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    write_mode = out_path.suffix.lower()
    if write_mode == ".xlsx":
        stats = update_xlsx(out_path, args.sheet, rows)
    elif write_mode == ".csv":
        with out_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            w.writeheader()
            w.writerows(rows)
        stats = None
    else:
        raise SystemExit(f"Unsupported output extension: {out_path.suffix} (use .csv or .xlsx)")

    # Console summary
    print()
    print(f"Series folders scanned : {scanned_series_count}")
    print(f"Collections matched    : {len(collections)}")
    print(f"Missing entries        : {len(rows)}")
    if stats is not None:
        print(f"XLSX merge             : kept {stats['kept']}, added {stats['added']}, "
              f"deleted {stats['deleted']} (sheet: '{args.sheet}')")
        print(f"Workbook               : {out_path}")
    else:
        print(f"CSV report             : {out_path}")
    if unmatched_series:
        print()
        print(f"{len(unmatched_series)} series folder(s) had no TMDB collection match")
        print("(these may be your own groupings rather than an official franchise):")
        for folder, n in unmatched_series:
            print(f"  - {folder}  ({n} movies)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
