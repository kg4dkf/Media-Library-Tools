"""Offline smoke test for find_missing_series_movies.py.

Builds a fake library on disk, mocks TMDB.get, runs main(), and checks the CSV.
"""
import csv
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))

import find_missing_series_movies as fm


# ---------------------------------------------------------------------------
# Fake TMDB responses keyed by URL path + sorted params
# ---------------------------------------------------------------------------
#
# Library we'll build:
#   /Matrix/
#     The Matrix 1999 imdb-tt0133093/file.mkv
#     The Matrix Reloaded 2003 imdb-tt0234215/file.mkv
#     # MISSING: The Matrix Revolutions (tt0242653)
#     # MISSING: The Matrix Resurrections (tt10838180)
#   /Lord of the Rings/
#     The Fellowship of the Ring 2001 imdb-tt0120737/file.mkv
#     The Two Towers 2002 imdb-tt0167261/file.mkv
#     The Return of the King 2003 imdb-tt0167260/file.mkv
#     # Nothing missing - all 3 present
#   /Director Spotlight - Nolan/   (user grouping, no TMDB collection)
#     Memento 2000 imdb-tt0209144/file.mkv
#   /The Shawshank Redemption 1994 imdb-tt0111161/file.mkv  (standalone movie, skipped)

FAKE_DB = {
    # /find/ttXXXX returns the movie's TMDB id
    "tt0133093": {"movie_results": [{"id": 603}]},
    "tt0234215": {"movie_results": [{"id": 604}]},
    "tt0120737": {"movie_results": [{"id": 120}]},
    "tt0167261": {"movie_results": [{"id": 121}]},
    "tt0167260": {"movie_results": [{"id": 122}]},
    "tt0209144": {"movie_results": [{"id": 77}]},
}

# /movie/{id} responses with collection info
MOVIE_DETAIL = {
    603: {"id": 603, "belongs_to_collection": {"id": 2344, "name": "The Matrix Collection"}},
    604: {"id": 604, "belongs_to_collection": {"id": 2344, "name": "The Matrix Collection"}},
    120: {"id": 120, "belongs_to_collection": {"id": 119, "name": "The Lord of the Rings Collection"}},
    121: {"id": 121, "belongs_to_collection": {"id": 119, "name": "The Lord of the Rings Collection"}},
    122: {"id": 122, "belongs_to_collection": {"id": 119, "name": "The Lord of the Rings Collection"}},
    77:  {"id": 77,  "belongs_to_collection": None},  # Memento - not in a TMDB collection
}

# /collection/{id} responses
COLLECTIONS = {
    2344: {
        "id": 2344, "name": "The Matrix Collection",
        "parts": [
            {"id": 603, "title": "The Matrix", "release_date": "1999-03-30"},
            {"id": 604, "title": "The Matrix Reloaded", "release_date": "2003-05-15"},
            {"id": 605, "title": "The Matrix Revolutions", "release_date": "2003-11-05"},
            {"id": 624860, "title": "The Matrix Resurrections", "release_date": "2021-12-16"},
        ],
    },
    119: {
        "id": 119, "name": "The Lord of the Rings Collection",
        "parts": [
            {"id": 120, "title": "The Fellowship of the Ring", "release_date": "2001-12-18"},
            {"id": 121, "title": "The Two Towers", "release_date": "2002-12-18"},
            {"id": 122, "title": "The Return of the King", "release_date": "2003-12-17"},
        ],
    },
}

# /movie/{id}/external_ids
EXTERNAL_IDS = {
    605: {"imdb_id": "tt0242653"},
    624860: {"imdb_id": "tt10838180"},
}


def fake_get(self, path, params=None):
    params = params or {}
    if path.startswith("/find/"):
        tt = path.split("/")[-1].lower()
        return FAKE_DB.get(tt)
    if path.startswith("/movie/") and path.endswith("/external_ids"):
        mid = int(path.split("/")[2])
        return EXTERNAL_IDS.get(mid, {"imdb_id": None})
    if path.startswith("/movie/"):
        mid = int(path.split("/")[-1])
        return MOVIE_DETAIL.get(mid)
    if path.startswith("/collection/"):
        cid = int(path.split("/")[-1])
        return COLLECTIONS.get(cid)
    raise AssertionError(f"unexpected TMDB call: {path} {params}")


def build_library(root: Path):
    def mk(p):
        p.mkdir(parents=True, exist_ok=True)
        (p / "dummy.mkv").write_text("x")

    matrix = root / "Matrix"
    mk(matrix / "The Matrix 1999 imdb-tt0133093")
    mk(matrix / "The Matrix Reloaded 2003 imdb-tt0234215")

    lotr = root / "Lord of the Rings"
    mk(lotr / "The Fellowship of the Ring 2001 imdb-tt0120737")
    mk(lotr / "The Two Towers 2002 imdb-tt0167261")
    mk(lotr / "The Return of the King 2003 imdb-tt0167260")

    nolan = root / "Director Spotlight - Nolan"
    mk(nolan / "Memento 2000 imdb-tt0209144")

    # Top-level standalone (should be skipped)
    mk(root / "The Shawshank Redemption 1994 imdb-tt0111161")


def main():
    with tempfile.TemporaryDirectory() as tdir:
        root = Path(tdir) / "Movies"
        root.mkdir()
        build_library(root)

        cache = Path(tdir) / "cache.json"
        out_csv = Path(tdir) / "missing.csv"

        os.environ["TMDB_API_KEY"] = "fake-key"

        with patch.object(fm.TMDB, "get", fake_get):
            rc = fm.main([str(root), str(out_csv), "--cache", str(cache), "--sleep", "0"])
            assert rc == 0, f"exit code {rc}"

        with out_csv.open(encoding="utf-8") as f:
            rows = list(csv.DictReader(f))

        # Expectations
        titles_by_collection = {}
        for r in rows:
            titles_by_collection.setdefault(r["collection_name"], []).append(r["missing_title"])

        # Matrix collection: 2 missing (Revolutions, Resurrections)
        matrix_missing = sorted(titles_by_collection.get("The Matrix Collection", []))
        assert matrix_missing == ["The Matrix Resurrections", "The Matrix Revolutions"], \
            f"Matrix missing wrong: {matrix_missing}"

        # LOTR: complete, no missing rows
        assert "The Lord of the Rings Collection" not in titles_by_collection, \
            "LOTR should have no missing entries"

        # Nolan grouping: should appear in unmatched (printed) and NOT in CSV
        assert all(r["collection_name"] != "Director Spotlight - Nolan" for r in rows)

        # IMDb id should be filled in via external_ids lookup
        revs = [r for r in rows if r["missing_title"] == "The Matrix Revolutions"][0]
        assert revs["missing_imdb_id"] == "tt0242653", revs
        assert revs["missing_year"] == "2003", revs

        print()
        print("CSV SMOKE TEST PASSED")
        print(f"  CSV rows: {len(rows)}")
        for r in rows:
            print(f"  - [{r['collection_name']}] {r['missing_title']} ({r['missing_year']})  {r['missing_imdb_id']}")

        # ------------------------------------------------------------------
        # XLSX merge test
        # ------------------------------------------------------------------
        from openpyxl import load_workbook

        xlsx_path = Path(tdir) / "report.xlsx"

        # First run: file doesn't exist - script should create it.
        with patch.object(fm.TMDB, "get", fake_get):
            rc = fm.main([str(root), str(xlsx_path),
                          "--sheet", "missing movies",
                          "--cache", str(cache), "--sleep", "0"])
            assert rc == 0

        wb = load_workbook(xlsx_path)
        assert "missing movies" in wb.sheetnames
        ws = wb["missing movies"]
        # header + 2 rows (Revolutions, Resurrections)
        assert ws.max_row == 3, f"expected 3 rows, got {ws.max_row}"
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "missing_imdb_id" in headers
        # Simulate the user adding a note column + a highlight on one row.
        notes_col = ws.max_column + 1
        ws.cell(row=1, column=notes_col, value="Notes")
        # Find the Resurrections row and tag it.
        imdb_col = headers.index("missing_imdb_id") + 1
        title_col = headers.index("missing_title") + 1
        resurrections_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=title_col).value == "The Matrix Resurrections":
                resurrections_row = r
                break
        assert resurrections_row is not None
        ws.cell(row=resurrections_row, column=notes_col, value="HIGH PRIORITY - 4K UHD")
        wb.save(xlsx_path)

        # Now: user "acquires" The Matrix Revolutions - add the folder.
        (root / "Matrix" / "The Matrix Revolutions 2003 imdb-tt0242653").mkdir()
        (root / "Matrix" / "The Matrix Revolutions 2003 imdb-tt0242653" / "x.mkv").write_text("x")
        # Pretend TMDB knows about it now:
        FAKE_DB["tt0242653"] = {"movie_results": [{"id": 605}]}
        MOVIE_DETAIL[605] = {"id": 605, "belongs_to_collection": {"id": 2344, "name": "The Matrix Collection"}}

        # Second run: should DELETE Revolutions row, KEEP Resurrections row (and its note).
        with patch.object(fm.TMDB, "get", fake_get):
            rc = fm.main([str(root), str(xlsx_path),
                          "--sheet", "missing movies",
                          "--cache", str(cache), "--sleep", "0"])
            assert rc == 0

        wb = load_workbook(xlsx_path)
        ws = wb["missing movies"]
        # header + 1 row (Resurrections only)
        assert ws.max_row == 2, f"expected 2 rows after acquisition, got {ws.max_row}"

        # Re-resolve column indices (in case user-added Notes col is preserved).
        headers2 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        title_col2 = headers2.index("missing_title") + 1
        notes_col2 = headers2.index("Notes") + 1  # MUST still exist
        assert ws.cell(row=2, column=title_col2).value == "The Matrix Resurrections"
        assert ws.cell(row=2, column=notes_col2).value == "HIGH PRIORITY - 4K UHD", \
            "user's Note must be preserved across the merge"

        print()
        print("XLSX MERGE SMOKE TEST PASSED")
        print(f"  Sheet rows after merge: {ws.max_row - 1}")
        print(f"  User 'Notes' column preserved with text: "
              f"{ws.cell(row=2, column=notes_col2).value!r}")

        # ------------------------------------------------------------------
        # Top-level credit test
        # ------------------------------------------------------------------
        # Move the user's "acquired" Resurrections to the TOP LEVEL of H:\Movies
        # instead of inside the Matrix series folder. The script should still
        # credit it because Matrix Collection is tracked via the series folder.
        import shutil
        top_level_rebirth = root / "The Matrix Resurrections 2021 imdb-tt10838180"
        top_level_rebirth.mkdir()
        (top_level_rebirth / "x.mkv").write_text("x")
        # Need TMDB to know about it so the parts list lookup pre-populates the
        # imdb_to_tracked map. tt10838180 wasn't in FAKE_DB yet.
        FAKE_DB["tt10838180"] = {"movie_results": [{"id": 624860}]}
        MOVIE_DETAIL[624860] = {"id": 624860, "belongs_to_collection": {"id": 2344, "name": "The Matrix Collection"}}

        xlsx2 = Path(tdir) / "report_top_level.xlsx"
        with patch.object(fm.TMDB, "get", fake_get):
            rc = fm.main([str(root), str(xlsx2),
                          "--sheet", "missing movies",
                          "--cache", str(Path(tdir) / "cache2.json"),
                          "--sleep", "0"])
            assert rc == 0

        wb = load_workbook(xlsx2)
        ws = wb["missing movies"]
        # All Matrix collection parts now present (Matrix + Reloaded + Revolutions
        # in the Matrix series folder; Resurrections at top level). LOTR complete
        # already. Should be 0 missing.
        rows_after = [
            [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1)
        ]
        rows_after = [r for r in rows_after if any(v not in (None, "") for v in r)]
        assert rows_after == [], f"expected no missing rows, got: {rows_after}"

        print()
        print("TOP-LEVEL CREDIT SMOKE TEST PASSED")
        print("  Top-level Resurrections correctly credited against Matrix Collection")


if __name__ == "__main__":
    main()
