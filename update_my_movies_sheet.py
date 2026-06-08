#!/usr/bin/env python3
"""
update_my_movies_sheet.py — Refresh the 'my_movies' sheet inside an
existing xlsx workbook from a CSV produced by library_export.py.

Why this script (rather than letting library_export.py write xlsx
directly): the workbook contains a summary sheet whose formulas
reference the my_movies sheet. Deleting and recreating the sheet
would break those references. Instead we clear the data rows under
the header and rewrite them in place.

What gets preserved:
  - All other sheets (summary, etc.) and their formulas
  - The existing header cells of the target sheet (kept in place; new
    columns are only ever appended at the end, never inserted in the
    middle, so summary-sheet formulas referencing existing columns
    don't shift)
  - Any extra columns in the target sheet that aren't in the CSV
    (e.g., a personal 'watched' column you might have added)
  - Cell formatting on rows we clear (only values are blanked)

What gets overwritten:
  - The cell VALUES under the headers that match CSV column names

What may get APPENDED to the header row (one-time, on first run):
  - 'Length (min)' — if the sheet has no column for runtime/length/
    duration and the enriched CSV contains a runtime_min column. After
    the first run the script just keeps populating the same column.

Usage:
    py update_my_movies_sheet.py \
        --csv  "C:\\Users\\me\\OneDrive\\Documents\\my_movies.csv" \
        --xlsx "H:\\_mediaprep\\series_stats_details.xlsx" \
        --sheet my_movies
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install it with:", file=sys.stderr)
    print("    py -m pip install openpyxl", file=sys.stderr)
    sys.exit(2)


VERSION = "1.0.0"


def main() -> int:
    ap = argparse.ArgumentParser(prog="update_my_movies_sheet.py",
                                 description=__doc__.splitlines()[1])
    ap.add_argument("--csv", required=True,
                    help="source CSV (produced by library_export.py)")
    ap.add_argument("--xlsx", required=True,
                    help="target workbook (.xlsx)")
    ap.add_argument("--sheet", default="my_movies",
                    help="target sheet name (default: my_movies)")
    args = ap.parse_args()

    csv_path = Path(args.csv)
    xlsx_path = Path(args.xlsx)
    if not csv_path.exists():
        print(f"ERROR: CSV not found: {csv_path}", file=sys.stderr)
        return 3
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}", file=sys.stderr)
        return 3

    # --- Load CSV first so we fail fast on malformed input -----------
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        csv_headers = list(reader.fieldnames or [])
        if not csv_headers:
            print("ERROR: CSV has no header row", file=sys.stderr)
            return 4
        csv_rows = list(reader)
    print(f"Loaded {len(csv_rows)} rows from {csv_path.name}")

    # --- Open the workbook (keeps every other sheet and its formulas) -
    try:
        wb = load_workbook(xlsx_path)
    except PermissionError:
        print(f"ERROR: Could not open {xlsx_path}.", file=sys.stderr)
        print("       Is the workbook open in Excel? Close it and retry.",
              file=sys.stderr)
        return 5

    if args.sheet not in wb.sheetnames:
        print(f"ERROR: sheet '{args.sheet}' not found in workbook.",
              file=sys.stderr)
        print(f"       Available sheets: {', '.join(wb.sheetnames)}",
              file=sys.stderr)
        return 6
    ws = wb[args.sheet]

    # --- Match CSV columns to sheet columns by header name -----------
    sheet_headers = [str(cell.value or "").strip() for cell in ws[1]]
    if not any(sheet_headers):
        print(f"ERROR: sheet '{args.sheet}' has no header row. "
              "Refusing to write to avoid clobbering data.", file=sys.stderr)
        return 7
    header_idx = {h.lower(): i + 1 for i, h in enumerate(sheet_headers) if h}

    # Some CSV columns are "aliased" — they have several acceptable
    # sheet-header names, listed in priority order. Aliased columns are
    # always resolved through the alias list (NOT direct-name match),
    # so a user-added 'Length' column wins over a leftover 'runtime_min'
    # column when both are present in the sheet.
    ALIAS_GROUPS = [
        # canonical csv name, default header to add, list of accepted
        # sheet headers (in priority order — first match wins)
        ("runtime_min", "Length (min)",
         ["length", "length (min)", "length_min", "length min",
          "runtime", "runtime (min)", "runtime_min", "runtime min",
          "duration", "duration_min", "duration (min)"]),
    ]
    aliased_csv = {canonical for canonical, _, _ in ALIAS_GROUPS}

    # Pass 1 — direct header-name matches (skipping aliased columns; they
    # get their own resolution below).
    used_cols = []
    unmapped = []
    for h in csv_headers:
        if h.lower() in aliased_csv:
            continue
        if h.lower() in header_idx:
            used_cols.append((h, header_idx[h.lower()]))
        else:
            unmapped.append(h)

    # Pass 2 — alias resolution for known canonical columns.
    csv_keys_lower = {c.lower() for c in csv_headers}
    for canonical, default_header, aliases in ALIAS_GROUPS:
        if canonical not in csv_keys_lower:
            continue
        # Find every alias present in the sheet; pick the highest priority.
        matches = [(a, header_idx[a]) for a in aliases if a in header_idx]
        if matches:
            chosen_name, chosen_col = matches[0]
            chosen_display = sheet_headers[chosen_col - 1]
            print(f"Mapping CSV '{canonical}' -> sheet column "
                  f"'{chosen_display}'")
            if len(matches) > 1:
                others = [sheet_headers[c - 1] for _, c in matches[1:]]
                print(f"  NOTE: sheet has multiple runtime-like columns: "
                      f"{[chosen_display] + others}. Using "
                      f"'{chosen_display}'; the others are left untouched "
                      f"(consider deleting them to avoid confusion).",
                      file=sys.stderr)
        else:
            # Append a new column at the end of the existing header row.
            chosen_col = len(sheet_headers) + 1
            ws.cell(row=1, column=chosen_col).value = default_header
            sheet_headers.append(default_header)
            header_idx[default_header.lower()] = chosen_col
            print(f"Added new sheet column '{default_header}' at position "
                  f"{chosen_col}")
        used_cols.append((canonical, chosen_col))

    if unmapped:
        print(f"WARNING: CSV columns not present in sheet (skipping): "
              f"{unmapped}", file=sys.stderr)
    if not used_cols:
        print("ERROR: No CSV columns matched sheet headers. "
              "Check the sheet's row 1 names.", file=sys.stderr)
        return 8

    # --- Clear existing data rows (keep formatting; preserve unrelated
    # columns the user may have added like a personal 'watched' flag) --
    max_row = ws.max_row
    cleared = 0
    if max_row >= 2:
        for row in range(2, max_row + 1):
            for _, col in used_cols:
                ws.cell(row=row, column=col).value = None
            cleared += 1
    print(f"Cleared {cleared} existing data row(s) "
          f"in '{args.sheet}' (header preserved)")

    # --- Write fresh CSV data starting at row 2 ----------------------
    # Numeric columns are coerced from strings to ints so Excel treats
    # them as numbers (sortable, comparable, formattable).
    NUMERIC_CSV_COLS = {"runtime_min"}
    for i, row in enumerate(csv_rows, start=2):
        for csv_col, sheet_col in used_cols:
            val = row.get(csv_col, "")
            if csv_col.lower() in NUMERIC_CSV_COLS and val:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass  # leave as string if non-numeric
            ws.cell(row=i, column=sheet_col).value = val if val != "" else None

    # --- Save --------------------------------------------------------
    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"ERROR: Could not save {xlsx_path}. Is the workbook open?",
              file=sys.stderr)
        return 5
    print(f"Wrote {len(csv_rows)} rows to {xlsx_path} :: {args.sheet}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
